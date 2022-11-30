from datetime import datetime
from http.client import BAD_REQUEST
from .model import Tratamentos, Usuario, Agendamento, Produtos
from ext import login_maneger, db
from flask import render_template, request, redirect, url_for, flash, session
from flask_login import UserMixin, login_required, login_user, logout_user
from werkzeug.security import generate_password_hash, check_password_hash
from flask_mail import Mail, Message
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def init_app(app):
    
    
    config = {
        "MAIL_SERVER": "smtp.ethereal.email",
        "MAIL_PORT": 587,
        "MAIL_USE_TLS": True,
        "MAIL_DEBUG": True,
        "MAIL_USERNAME": "katelyn.lubowitz1@ethereal.email",
        "MAIL_PASSWORD": "AHRpb2F4vZ5vZ1VUTE",
        "MAIL_DEFAULT_SENDER": "Espaço PB <espacopb@gmail.com.br>",
    }
    app.config.update(config)
    mail = Mail(app)
    
    @app.before_first_request
    def create_db():
        db.create_all()

    @login_maneger.user_loader
    def current_user(id):
        return Usuario.query.get(id)

    @app.route('/')
    def index():
        return render_template('index.html')

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'GET':
            return render_template('login-v2.html')
        elif request.method == 'POST':
            email = request.form['email']
            senha = request.form['senha']
            usuario = Usuario.query.filter_by(email=email).first()
            session['usuario'] = usuario.id
            if usuario and check_password_hash(usuario.senha, senha):
                login_user(usuario)
                if usuario.in_admin == 1:
                    return redirect (url_for('dash_admin', id=usuario.id))
                elif usuario.in_colaborador == 1:
                    return redirect (url_for('dash_colaborador', id=usuario.id))
                else:
                    return redirect(url_for('dash_cliente', id=usuario.id))
            else:
                flash('Credências Invalidas.')
                return redirect(url_for('login'))
        else:
            return BAD_REQUEST

    @app.route('/cadastrar', methods=['GET', 'POST'])
    def cadastrar():
        if request.method == 'GET':
            return render_template('cadastro.html')
        elif request.method == 'POST':
            user = Usuario()
            user.nome = request.form['nome']
            user.rua = request.form['rua']
            user.numero = request.form['numero']
            user.bairro = request.form['bairro']
            user.cidade = request.form['cidade']
            user.estado = request.form['estado']
            user.fone = request.form['fone']
            user.email = request.form['email']
            user.senha = generate_password_hash(request.form['senha'])
            user.in_admin = 0
            user.in_colaborador = 1
            user.in_ativo = 1
            
            db.session.add(user)
            db.session.commit()
            
            return redirect(url_for('login'))
        else:
            return BAD_REQUEST
    
    @app.route('/dashboard/<id>', methods=['GET', 'POST'])
    @login_required
    def dash_cliente(id):
        usuario = Usuario.query.filter_by(id = id).first()
        agendamentos = Agendamento.query.filter_by(id = id).count()
        if usuario.in_colaborador == 0:
            agendamentos =  Agendamento.query.filter_by(id_usuario = id).count()
            return render_template('dash_cliente/dashboard_cliente.html', usuario=usuario.nome, agendamentos=agendamentos)
        else:
            return BAD_REQUEST
        
    
    @app.route('/dashboard_colaborador/<id>', methods=['GET', 'POST'])
    @login_required
    def dash_colaborador(id):
        usuario = Usuario.query.filter_by(id = id).first()
        agendamentos = Agendamento.query.filter_by().count()
        if usuario.in_colaborador == 1:
            agendamentos =  Agendamento.query.count()
            agendamento_realizado = Agendamento.query.filter_by(in_realizado=1, id_usuario_finalizacao=id).count()
            porcentagem = (agendamento_realizado * 100) / agendamentos
            return render_template('dash_colaborador/dashboard_colaborador.html', agendamentos=agendamentos, porcentagem=round(porcentagem), colaborador=usuario.nome)
        else:
            return BAD_REQUEST
    
    @app.route('/dashboard_admin/<id>', methods=['GET', 'POST'])
    @login_required
    def dash_admin(id):
        usuario = Usuario.query.filter_by(id = id).first()
        agendamentos = Agendamento.query.filter_by().count()
        novos_usuarios = Usuario.query.filter_by().count()
        produtos = Tratamentos.query.count()
        tratamentos = Produtos.query.count()
        
        if usuario.in_admin == 1:
            return render_template('dash_admin/dashboard_admin.html',
                           novos_usuarios=novos_usuarios,
                           agendamentos=agendamentos,
                           usuarios=novos_usuarios,
                           tratamentos=tratamentos,
                           produtos=produtos)
        else:
            return BAD_REQUEST
 
    @app.route('/logout', methods=['GET', 'POST'])
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('index'))
    
    @app.route('/agendamento', methods=['GET', 'POST'])
    @login_required
    def agendamento():
        if request.method == 'GET':
            listar_produtos = Produtos.query.all()
            listar_tratamentos = Tratamentos.query.all()
            listar_profissionais = Usuario.query.filter_by(in_colaborador = 1).all()
            return render_template('dash_cliente/agendamentos_cliente.html', listar_produtos=listar_produtos, listar_profissionais=listar_profissionais, listar_tratamentos=listar_tratamentos)
        elif request.method == 'POST':
            agendamento = Agendamento()
            agendamento.id_usuario = session['usuario']
            agendamento.nome = f'{request.form["nome"]} {request.form["sobrenome"]}'
            agendamento.email = request.form['email']
            agendamento.fone = request.form['telefone']
            agendamento.data_agendamento = request.form['dt_agendamento']
            agendamento.horario_agendamento = request.form['hr_agendamento']
            agendamento.colaborador = request.values.get('cabeleleiro')
            agendamento.data_registro = datetime.now()
            agendamento.cd_tratamento = request.values.get('tratamentos')
            
            info_tratamento = Tratamentos.query.filter_by(id=2).all()
            agendamento.valor_tratamento = info_tratamento[0].valor
            agendamento.valor_produto = 0
            
            
            db.session.add(agendamento)
            db.session.commit()
            usuario = Usuario.query.filter_by(in_colaborador=1).first()
            usuario_admin = Usuario.query.filter_by(in_admin=1).first()
            sendmail(assunto='Agendamento solicitado', email=agendamento.email, body=f'Olá, Seu agendamento para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento} foi realizado. \n Aguarde a confirmação')
            sendmail(assunto='Agendamento solicitado', email=usuario.email, body=f'Olá, O cliente {agendamento.nome} solicitou um Agendamento para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento}. \n Caso tenha disponibilidade confirme o agendamento,')
            sendmail(assunto='Agendamento solicitado', email=usuario_admin.email, body=f'Olá, O cliente {agendamento.nome} solicitou um Agendamento para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento}. \n Caso tenha disponibilidade confirme o agendamento,')
            return render_template('dash_cliente/agendamentos_cliente.html')
        else:
            return BAD_REQUEST
     
        
    @app.route('/agendamento_editar/<id>', methods=['GET', 'POST'])
    @login_required
    def agendamento_editar(id):
        usuario = session['usuario']
        if request.method == 'GET':
            agendamento = Agendamento.query.filter_by(id=id).all()
            return render_template('dash_colaborador/agendamentos_colaborador.html', agendamento=agendamento)
        elif request.method == 'POST':
            agendamento = Agendamento.query.filter_by(id=id).first()
            usuario = Usuario.query.filter_by(id=usuario).first()
            # agendamento.id_usuario = session['usuario']
            agendamento.nome = request.form["nome"]
            agendamento.email = request.form['email']
            agendamento.fone = request.form['telefone']
            agendamento.data_agendamento = request.form['dt_agendamento']
            agendamento.horario_agendamento = request.form['hr_agendamento']
            agendamento.colaborador = request.form['cabeleleiro']
            agendamento.in_confimado = ( 1 if request.form['confirmado'] else 0)
            if agendamento.in_confimado == 1:
                agendamento.id_usuario_confirma = usuario.id
            agendamento.data_registro = datetime.now()
            agendamento.cd_tratamento = request.values.get('tratamentos')
            agendamento.cd_produto = request.values.get('produtos')
            
            db.session.add(agendamento)
            db.session.commit()
            usuario = Usuario.query.filter_by(in_colaborador=1).first()
            usuario_admin = Usuario.query.filter_by(in_admin=1).first()
            sendmail(assunto='Agendamento Confirmado', email=agendamento.email, body=f'Olá, Seu agendamento para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento} foi Confirmado. \n o Procedimento será realizado pelo profissional: {usuario.nome}')
            sendmail(assunto='Agendamento Confirmado', email=usuario_admin.email, body=f'Olá, Foi confirmado o agendamento com o cliente {agendamento.nome} para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento}. \n Após o atendimento finalize o agendamento,')
            sendmail(assunto='Agendamento Confirmado', email=usuario.email, body=f'Olá, Foi confirmado o agendamento com o cliente {agendamento.nome} para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento}. \n Após o atendimento finalize o agendamento,')
            
            return render_template('dash_colaborador/agendamentos_colaborador.html')
        else:
            return BAD_REQUEST
        
    @app.route('/agendamento_finalizado/<id>', methods=['GET'])
    @login_required
    def agendamento_finalizado(id):
        usuario = session['usuario']
        agendamento = Agendamento.query.filter_by(id=id).first()
        agendamento.in_realizado = 1
        agendamento.id_usuario_finalizacao = usuario
         
        db.session.add(agendamento)
        db.session.commit()
        id_usuario = session['usuario']
        usuario = Usuario.query.filter_by(in_colaborador=1).first()
        usuario_admin = Usuario.query.filter_by(in_admin=1).first()
        sendmail(assunto='Agendamento Realizado', email=agendamento.email, body=f'Olá, Seu agendamento para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento} foi Realizado. \n o Processidimento foi realizado pelo profissional: {usuario.nome}')
        sendmail(assunto='Agendamento Realizado', email=usuario.email, body=f'Olá, O seu atendimento com o cliente {agendamento.nome} para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento} foi realizado. \n Obrigado!,')
        sendmail(assunto='Agendamento Realizado', email=usuario_admin.email, body=f'Olá, O atendimento com o cliente {agendamento.nome} para o dia: {agendamento.data_agendamento}, e horario: {agendamento.horario_agendamento}foi realizado, pelo profissional {usuario.nome}')
        return redirect(url_for('dash_colaborador', id=id_usuario))
         

    @app.route('/lista_agendamentos', methods=['GET', 'POST'])
    @login_required
    def lista_agendamentos():
        if 'usuario' in session:
            id_usuario = session['usuario']
            usuario = Usuario.query.filter_by(id = id_usuario).first()
            if usuario.in_admin == 1:
                valor_total_trat = 0
                valor_total_prod = 0
                lista_agendamentos = Agendamento.query.all()
                x = 0
                while len(lista_agendamentos) < x:
                    if lista_agendamentos[x].id_usuario_produtos:
                        lista_produtos = Produtos.query.filter_by(id=lista_agendamentos[x].id_usuario_produtos).all()
                        for xprod in lista_produtos:
                            valor_total_prod += xprod.valor
                    if lista_agendamentos[x].id_usuario_tratamentos:
                        lista_tratamentos = Tratamentos.query.filter_by(id=lista_agendamentos[x].id_usuario_tratamentos).all()
                        for xtrat in lista_tratamentos:
                            valor_total_trat += xtrat.valor
                    x += 1
                return render_template('dash_admin/listar_agendamentos_admin.html', lista_agendamentos=lista_agendamentos, valor_total=(valor_total_prod + valor_total_trat))
            elif usuario.in_colaborador == 1:
                lista_agendamentos = Agendamento.query.filter_by(id_usuario_finalizacao = None).all()
                return render_template('dash_colaborador/listar_agendamentos_colaborador.html', lista_agendamentos=lista_agendamentos)
            else:
                lista_agendamentos = Agendamento.query.filter_by(id_usuario=id_usuario).all()
                return render_template('dash_cliente/lista_agendamentos_cliente.html', lista_agendamentos=lista_agendamentos)
    
    @app.route('/lista_usuarios', methods=['GET', 'POST'])
    @login_required
    def lista_usuarios():
        if request.method == 'GET':
            usuarios = Usuario.query.all()
            return render_template('dash_admin/listar_usuarios.html', usuarios=usuarios)
        elif request.method == 'POST':
            usuario = Usuario.query.all()
            usuario.in_colaborador =  ( 1 if request.form['in_colaborador'] else 0)
            usuario.in_admin = ( 1 if request.form['in_admin'] else 0)
            usuario.in_ativo = ( 1 if request.form['in_ativo'] else 0)
            
            db.session.add(usuario)
            db.session.commit()
            id_usuario = session['usuario']
            return redirect(url_for('dash_admin', id=id_usuario))
    
    @app.route('/criar_produto', methods=['GET', 'POST'])
    @login_required
    def criar_produto():
        if request.method == 'GET':
            if 'usuario' in session:
                id_usuario = session['usuario']
                usuario = Usuario.query.filter_by(id = id_usuario).first()
                if usuario.in_admin == 1:
                    return render_template('dash_admin/cadastro_produto.html')
        elif request.method == 'POST':
            produto = Produtos()
            produto.nome = request.form['nome']
            produto.tipo = request.form['tipo']
            produto.valor = request.form['valor']
            produto.porcentagem = request.form['porcentagem']
            produto.id_usuario_criacao = session['usuario']
            produto.data_criacao = datetime.now()
            
            db.session.add(produto)
            db.session.commit()
            flash('Produto criado.')
            return render_template('dash_admin/cadastro_produto.html')

    @app.route('/criar_tratamento', methods=['GET', 'POST'])
    @login_required
    def criar_tratamento():
        if request.method == 'GET':
            if 'usuario' in session:
                id_usuario = session['usuario']
                usuario = Usuario.query.filter_by(id = id_usuario).first()
                if usuario.in_admin == 1:
                    return render_template('dash_admin/cadastro_tratamento.html')
        elif request.method == 'POST':
            tratamento = Tratamentos()
            tratamento.nome = request.form['nome']
            tratamento.tipo = request.form['tipo']
            tratamento.valor = request.form['valor']
            tratamento.porcentagem = request.form['porcentagem']
            tratamento.id_usuario_criacao = session['usuario']
            tratamento.data_criacao = datetime.now()
            
            db.session.add(tratamento)
            db.session.commit()
            flash('Tratamento criado.')
            return render_template('dash_admin/cadastro_tratamento.html')
        
    @app.route('/lista_produtos', methods=['GET'])
    @login_required
    def lista_produtos():
        if request.method == 'GET':
            produtos = Produtos.query.all()
            return render_template('dash_admin/listar_produtos.html', produtos=produtos)
        
    @app.route('/lista_tratamentos', methods=['GET'])
    @login_required
    def lista_tratamentos():
        if request.method == 'GET':
            tratamentos = Tratamentos.query.all()
            return render_template('dash_admin/listar_tratamentos.html', tratamentos=tratamentos)
    
    @app.route('/sendmail/<assunto>/<email>/<body>')
    @login_required
    def sendmail(assunto, email, body):
        msg = Message(
            subject= assunto, 
            sender= app.config["MAIL_DEFAULT_SENDER"], 
            recipients=[email], 
            body=body
            )
        mail.send(msg)
        return 'email enviado'


    @app.route('/relatorio_gerencial_admin', methods=['GET'])
    @login_required
    def rel_ger_admin():
        if request.method == 'GET':
            agendamento = Agendamento.query.all()
            print(agendamento)
            wb = Workbook()

            dest_filename = 'rel_geral_admin.xlsx'

            ws1 = wb.active
            ws1.title = "Agendamentos"

            ws1.append(['Nome', 'Email', 'Data Agendamento', 'Valor Produto', 'Valor Tratamento'])
            
            for x in agendamento:
                print(x.nome)
                ws1.append([x.nome, x.email, x.data_agendamento, x.valor_produto, x.valor_tratamento])

            wb.save(filename = dest_filename)
        return ''
    
    @app.route('/relatorio_gerencial_admin_por_colab', methods=['GET'])
    @login_required
    def relatorio_gerencial_admin_por_colab():
        if request.method == 'GET':
            agendamento = Agendamento.query.all()
            print(agendamento)
            wb = Workbook()
            dest_filename = 'relatorio_gerencial_admin_por_colab.xlsx'
            ws1 = wb.active
            ws1.title = "Realizados_por_colab"
            ws1.append(['Nome','realizado', 'Data Agendamento', 'Valor venda Produto', 'Valor venda Tratamento'])
            
            for x in agendamento:
                print(x.nome)
                colab = Usuario.query.filter_by(id=x.colaborador).first()
                produtos = Produtos.query.filter_by(id = x.cd_produto).all()
                tratamentos = Tratamentos.query.filter_by(id = x.cd_tratamento).all()
                ws1.append([colab.nome, ('Sim' if x.in_realizado == 1 else 'Não'), x.data_agendamento, (produtos[0].valor * (int(produtos[0].porcentagem) / 100) if produtos[0].valor else 0), (tratamentos[0].valor * (int(tratamentos[0].porcentagem) / 100) if tratamentos[0].valor else 0)])
            wb.save(filename = dest_filename)
        return ''